import random

# for i in range(3):
#     print(random.random())
#
#
# for i in range(3):
#     print(random.randint(10, 30))
#
#
# team = ["ali", "umair", "Usman", "hamza"]
# leader = random.choice(team)
# print(leader)


# class Dice:
#     def role(self):
#         first = random.randint(1,6)
#         second = random.randint(1,6)
#         return (first, second)
#
#
# dice1 = Dice()
# print(dice1.role())


from pathlib import Path
#absoulte path
#c:\programs\user\....
#/usr/local/bin....

#Relative Path

# path1 = Path("learning")
# print(path1.exists())

# # For creating New directory and Remove
# path1 = Path("address")
# print(path1.mkdir())
# print(path1.rmdir()) #Remove Directory

# #to Search all directory
# path1 = Path()
# for file in path1.glob("*"):
#     print(file)


### Automation With Python
import openpyxl as xl
from openpyxl.chart import Reference, BarChart

def process_spreadsheet(filename):
    book = xl.open(filename)
    sheet = book["Sheet1"]
    cell = sheet["a1"]
    # cell = sheet.cell(1,1)
    # print(cell.value) #to print value of cell
    # print(sheet.max_row) #to Find the number of rows in sheet
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        updated_price_cell = sheet.cell(row, 4)
        updated_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2, max_row=sheet.max_row,
                       min_col=4, max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "f3")

    book.save("transaction2.xlsx")


